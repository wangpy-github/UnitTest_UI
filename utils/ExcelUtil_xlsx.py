import os
from openpyxl import load_workbook

"""
openpyxl说明：
    可以实现对xlsx的读写操作，缺点是修改文件保存会丢失公式
"""


# 自定义异常
class SheetTypeError:
    pass


class ExcelReader_xlsx():
    def __init__(self, filepath, sheet_by_name):
        if os.path.exists(filepath):
            self.filepath = filepath
            self.sheet_by = sheet_by_name
            self._data = list()
        else:
            raise FileNotFoundError("Excel文件不存在")

    def data(self):
        title_list = list()
        workbook = load_workbook(self.filepath)
        sheet = workbook.get_sheet_by_name(self.sheet_by)
        # sheet.max_row 行数    sheet.max_column 列数
        # sheet['A1'].value  /   sheet.cell(row=1, column=1).value  单元格的值

        # 获取第一行的值列表
        for col in range(1, sheet.max_column + 1):
            title = sheet.cell(row=1, column=col).value
            title_list.append(title)

        # 获取每一行的值列表 组合数据，格式：[{"title1":"value1", "title2":"value2"},{"title1":"value3", "title2":"value4"} ]
        for row in range(1, sheet.max_column + 1):
            L = list()
            for col in range(1, sheet.max_column + 1):
                title = sheet.cell(row=row + 1, column=col).value
                L.append(title)
            self._data.append(dict(zip(title_list, L)))
        return self._data


if __name__ == '__main__':
    a = ExcelReader_xlsx("../data/test_data.xlsx", "帖子").data()
    print(a)
